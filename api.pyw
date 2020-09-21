import requests
import dotenv
import json
import pandas as pd
import openpyxl as xl

import time


class constants:
    name=0
    company_number=1
    Jurisdiction_code=2
    cmpnyType=3
    registry_url=4
    branch=5
    branch_status=6
    current_status=7
    street_add=8
    locality=9
    region=10
    postal_code=11
    country=12

const_list=["name","company_number","jurisdiction_code", "company_type", "registry_url", "branch","branch_status","current_status","street_address", "locality","region","postal_code","country" ]



# wb=Workbook()
# sheet1=wb.add_sheet("Sheet 1")
headigs=["Name", "Company Number", "Jurisdiction Code", "Company Type", "Registry URL", "Branch", "Branch Status", "Current Status", "Street Address", "Locality", "Region","Postal Code","Country"]

api_token=""
parameters={
    "industry_codes":"eu_nace_2-7211",
    "api_token":api_token,
    "per_page":100,
    "page":1
}
base_url="https://api.opencorporates.com/v0.4/"
companies_search_url="companies/search"
base_file_name="eu_nace_2-7211"
file_extn=".json"

def get_file_name(page):
    end=page*parameters['per_page']
    start=end-(parameters['per_page']-1)
    return base_file_name+"("+str(start)+"-"+str(end)+")"+file_extn

def create_files(page,text):
    name=get_file_name(page)
    
    try:
        print ('Creating file - ', name)
        # x=input('enter- ')
        with open(name,"w") as file:
            json.dump(text, file)
        file.close()
        print ('Created...')
        return True

    except Exception:
        return False
    

def create_excel():

    wb=xl.Workbook()
    sheet=wb.create_sheet("sheet1")
    for i in range(len(headigs)):
        sheet.cell(1, i+1, headigs[i])

    wb.save("excel1.xlsx")




def excel(dict1, page):
    
    # keys=dict1.keys()
    # items=dict1.items()
    #print ("keys - ", keys)
    #print ("items - ", items)
    # for i in range(len(headigs)):
    #     sheet1.write(0, i+1, headigs[i])

    # wb.save("excel.xls")
    wb=xl.load_workbook("excel1.xlsx")
    sheet1=wb.get_sheet_by_name("sheet1")

    try:
        k=0
        for j in range(page*parameters['per_page']-(parameters['per_page']-1)-1, page*parameters['per_page']):
            #for k in range(len(dict1['results']['companies'])):
            #print (dict1['results']['companies'])
            for i in range(len(const_list)):
                
                if(i<8):
                    print ('k=', k)
                    try:

                        # print (dict1['results']['companies'][j]['company'][const_list[i]])
                        # x=input("enter - ")

                    #print (dict1['results']['companies'][k]['company'][const_list[i]])
                    #print("j+2 - ", j+2, "i+1 - ", i+1, "k - ", k)
                    #x=input('enter - ')
                        sheet1.cell(j+2, i+1, dict1['results']['companies'][k]['company'][const_list[i]])

                    except IndexError:
                        print ("Index Error")
                        pass
                    
                if(i>=8 and i<=12):
                    
                    try:

                        #print (dict1['results']['companies'][j]['company']['registered_address'])
                        #print("j+2 - ", j+2, "i+1 - ", i+1, "k - ", k)

                        #x=input("enter - ")
                        sheet1.cell(j+2, i+1, dict1['results']['companies'][k]['company']['registered_address'][const_list[i]])
                    except TypeError:
                        print ("TypeError")
                        sheet1.cell(j+2, i+1, "Information not available")
                    except IndexError:
                        print ("Index Error")
                        pass
            k+=1
                

        wb.save("excel1.xlsx")

    except EOFError:
        print ("EOF Error")
        return -2
    except IndexError:
        print ("Index Error")
        return 0
    # except Exception as e:
    #     return -1
    return 0
    
def write_toExcel(page):

    name=get_file_name(page)

    with open(name, encoding="utf8") as file:
        print ("loading from "+ name)
        data=json.load(file) 
        # print (data) 
        file.close()

    return excel(json.loads(data, encoding="utf8"), page)
    
    
def get_companies(parameter, page):
    # page=0
    # endPage=92
    # parameters["page"]=page+1
    # while(page<endPage):
    try:

        response=requests.get(base_url+companies_search_url, params=parameter)
        text = json.dumps(response.json(), indent=0)
        return create_files(page, text)
    except TimeoutError:
        with open('erro pages.txt',  'a') as file:
            file.write(parameter['page']+"  ")
            return True



offline=2  #0 means offline, 1 means only online, 2 means do both

def loop_pages_api(off):
    page=91
    endPage=92
    while(page<endPage):
        
        parameters['page']=page+1
        print ("page in loop - ", page)
        print("parameters in loop - ", parameters['page'])
        if (off==1 or off==2):
            if not get_companies(parameters, parameters['page']):
                return False
            else: print ("get companies done")
            if(off==2): off=0
            
        if(off==0):
            print ("page - ", parameters['page'])
            print (write_toExcel(parameters['page']))
            off=offline
            #time.sleep(3)
        
        page+=1
        print ("page in loop end - ", page)
    return True


print (loop_pages_api(offline))
# response=requests.get(base_url+companies_search_url, params=parameters)
# text = json.dumps(response.json(), indent=0)


# with open(file_name,"w") as file:
#        json.dump(text, file)

# file.close()

# # # print ("73 - ", text[73])
# # # print ("74 - ", text[74])
# # # print ("75 - ", text[75])
# # # print ("context - ", text[40:100])
# # # # print ("to remove - ", text[4595:6615])
# # # print ("full text - ", text)
# # # text.replace(' ', '')
# # # print ("replaced text - ",text)




# with open(file_name, encoding="utf8") as file:
#    data=json.load(file) 
#    print (data) 


# dict2=json.loads(data, encoding="utf8")
# # # print (dict2['results']['companies'][1]['company']['name'])
# # """print ("dict2 - ", dict2)"""

# excel(dict2)
# # create_excel()

# dict1=pd.read_json(r'eu_nace_2-7211(1-20).json')
# dict1.to_csv(r'excel.csv', index=None)

# class Webapi:
#     url=""
#     parameters=None
#     def __init__(self, url, parameters):
#         self.url=url
#         self.parameters=parameters

#     def getCompanies(self,url, parameters):
#         response=requests.get(url, params=parameters)
#         json_response = json.dumps(response.json(), sort_keys=False, indent=4)
#         return json_response
    

# class Localdb:
#     file=None
#     def __init__(self, file):
#         self.file=file
    
#     def write_json(self, text):
#         with open('eu_nace_2-7211(1-20).json', 'w') as file:
#             json.dump(text, file)   
#             file.close()


# class Repository:
#     def __init__(self):
#         pass



# n=1
# if(n==1):
#     api=new Webapi(base_url+companies_search_url)
#     response=api.getCompanies()
#      with open('eu_nace_2-7211(1-20).json', 'w') as file:
#             json.dump(text, file)   
#             file.close()

# text=""
# with open("eu_nace_2-7211(1-20).json", "r") as file:
#     text=json.load(file)
#     file.close()

# dict=json.loads(text)
# print (dict)







